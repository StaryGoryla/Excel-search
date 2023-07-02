import openpyxl as xl

wb1 = xl.load_workbook("wingspan-20221201.xlsx")
sheet = wb1["Birds"]


def percent_of_all(x):
    percent = round(x / (sheet.max_row - 1) * 100)
    return percent


def sum_hits_in_column(x):
    sum = 0
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, x)
        if cell.value != None:
                sum += 1
    cell = sheet.cell(1, x)
    name = cell.value
    return sum, name


def sum_sole_hits_in_columns(x, y):
    sums = []
    names = []
    sum = 0
    continents = 0
    for column in range(x, y):
        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row, column)
            if cell.value != None:
                for temp_column in range(x, y):
                    cell2 = sheet.cell(row, temp_column)
                    if cell2.value != None:
                        continents += 1
                if continents == 1:
                    sum += 1
                continents = 0
        sums.append(sum)
        cell3 = sheet.cell(1, column)
        names.append(cell3.value)
        sum = 0
    return sums, names


def sum_values_in_column(x):
    sums = []
    sums2 = []
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, x)
        sums.append(cell.value)
    names = list(set(sums))
    for value in names:
        sums2.append(sums.count(value))
    return sums2, names


def sum_hits_in_columns(x, y):
    sum_bird = 0
    sum = []
    sums2 = []
    for row in range(2, sheet.max_row + 1):
        for column in range(x, y):
            cell = sheet.cell(row, column)
            if cell.value != None:
                sum_bird += cell.value
        sum.append(sum_bird)
        sum_bird = 0
    names = list(set(sum))
    for value in names:
        sums2.append(sum.count(value))
    return sums2, names


